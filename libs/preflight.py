#Version 3/1/24 - Add .ColumnValsMatchRegex method
#Version 8/26/24 - Add additional methods (from client code Version 6/24/24)
#Version 8/27/24 - Consolidate tbl.df and df checks to CheckDataFrame class
#                  Add tests of all preflights in test_preflight.py
#Version 8/29/24 - Set default tbl=None in case using Class to check a df
#                  without Table instance and its tbl.df attribute

import pandas as pd
import os
from openpyxl import load_workbook
import re
import util
from error_handling import ErrorHandle

#Initialize logging in case needed based on errs IsLog attribute
import logging
logger = logging.getLogger(__name__)
"""
=========================================================================
This class checks a projfiles.tbl.df (Table class df attribute) structure   
and datavalues based on specified Table attributes such as a list of 
required columns and the .df's default index

In alternate usage, df can be passed as an argument directly -- bypassing
use of the Table class to check df instead of self.tbl.df

The class relies on default or custom error codes in ErrorCodes.xlsx
whose df is an attribute of the .errs ErrorHandle instance. If 
.IsCustomCodes is True, the .errs.Locn attribute is used to look up a
custom error code and message in the ErrorCodes.xlsx file. Otherwise, 
the preflight.py method name is used to look up the default error code..
=========================================================================
"""
class CheckDataFrame:
    def __init__(self, path_err_codes, tbl=None, IsCustomCodes=False, \
                 IsPrint=True, IsLog=False):
        """
        Initialize CheckDataFrame
        JDL 2/16/24; Modified 8/29/24 to set default tbl=None
        """
        self.tbl = tbl
        self.IsPrint = IsPrint
        self.IsLog = IsLog
        self.errs = ErrorHandle(path_err_codes, '', IsHandle=True, IsPrint=IsPrint,)

        #If enabled, will not override errs.Locn with function name for code lookup
        self.IsCustomCodes = IsCustomCodes 

    def ContainsRequiredCols(self, cols_req=None, df=None):
        """
        .tbl.df contains specified list of column names (True if so)
        JDL 2/16/24; Modified 8/27/24 to add df arg
        """
        #Enable custom error codes and set df with precedence to arg df if supplied
        if not self.IsCustomCodes: self.errs.Locn = util.current_fn()
        if df is None: df = self.tbl.df

        #Enable override list of required columns list
        if cols_req is None: cols_req = self.tbl.required_cols

        #Check df has all required columns
        for col in cols_req:
            if self.errs.is_fail((not col in df.columns), 1, self.errs.Locn, col):
                self.errs.RecordErr()
                return False
        return True
    
    def NoDuplicateCols(self, df=None):
        """
        .tbl.df has unique column names (True if so)
        JDL 2/16/24; Modified 8/27/24 to add df arg
        """
        #Enable custom error codes and set df with precedence to arg df if supplied
        if not self.IsCustomCodes: self.errs.Locn = util.current_fn()
        if df is None: df = self.tbl.df

        #Check if there are true duplicate column names
        cols = df.columns
        if not cols.is_unique: 
            duplicates = cols[cols.duplicated()].unique()
            ErrParam = '\nDuplicate columns: ' + ', '.join(map(str, duplicates))
            if self.errs.is_fail(True, 2, self.errs.Locn, ErrParam): self.errs.RecordErr()
            return False
        
        #Look for Pandas modified names from pd.read_excel (e.g. .1, .2, etc.)
        else:
            #Check if any column has Pandas-added extension (.1, .2, etc.)
            for col in cols:
                scol = str(col)
                if '.' in scol and scol.rsplit('.', 1)[1].isdigit():
                    ErrParam = scol.split('.')[0]
                    if self.errs.is_fail(True, 2, self.errs.Locn, ErrParam): self.errs.RecordErr()
                    return False
                
        #No duplicates detected
        return True

    def NoDuplicateIndices(self, df=None):
        """
        .tbl.df has unique index values (True if so)
        JDL 2/16/24; Modified 8/27/24 to add df arg
        """
        #Enable custom error codes and set df with precedence to arg df if supplied
        if not self.IsCustomCodes: self.errs.Locn = util.current_fn()
        if df is None: df = self.tbl.df

        idx = df.index
        if idx.is_unique: return True

        #Report duplicate index values
        duplicates = idx[idx.duplicated()].unique()
        ErrParam = '\nDuplicate indices: ' + ', '.join(map(str, duplicates))
        if self.errs.is_fail(True, 3, self.errs.Locn, ErrParam): self.errs.RecordErr()
        return False

    def LstColsPopulated(self, df=None, lst_cols=None):
        """
        .tbl.df list of tbl.populated_cols populated with non-blank values (True if so)
        JDL 2/16/24; Modified 8/27/24 to add df and lst_cols args
        """
        #list to check is tbl attribute unless overridden by args
        if lst_cols is None: lst_cols = self.tbl.populated_cols

        for col in lst_cols:
            if not self.ColPopulated(col, df=df): return False
        return True

    def ColPopulated(self, col_name, df=None):
        """
        All values in a specified column are non-null (True if so)
        JDL 2/16/24; Modified 8/27/24 to add df arg
        """
        #Enable custom error codes and set df with precedence to arg df if supplied
        if not self.IsCustomCodes: self.errs.Locn = util.current_fn()
        if df is None: df = self.tbl.df

        #Report blank values in specified column
        if not df[col_name].isna().any(): return True
        if self.errs.is_fail(True, 4, self.errs.Locn, col_name): self.errs.RecordErr()
        return False

    def ColumnsContainListVals(self, list_vals, df=None):
        """
        DataFrame columns contain a specified list of values
        JDL 2/16/24; Modified 8/27/24 to add df arg
        """
        #Enable custom error codes and set df with precedence to arg df if supplied
        if not self.IsCustomCodes: self.errs.Locn = util.current_fn()
        if df is None: df = self.tbl.df

        # Loop over list_vals and check if in DataFrame columns
        for val in list_vals:
            if val not in df.columns:
                ErrParam = '\nMissing: ' + str(val)
                if self.errs.is_fail(True, 5, self.errs.Locn, ErrParam): self.errs.RecordErr()
                return False
        return True
    
    def IndexContainsListVals(self, list_vals, df=None):
        """
        DataFrame index contains a specified list of values
        JDL 1/11/24; Modified 8/27/24 to add df arg
        """
        #Enable custom error codes and set df with precedence to arg df if supplied
        if not self.IsCustomCodes: self.errs.Locn = util.current_fn()
        if df is None: df = self.tbl.df

        # Loop over list_vals and check if in DataFrame index
        for val in list_vals:
            if val not in df.index:
                ErrParam = '\nMissing: ' + str(val)
                if self.errs.is_fail(True, 6, self.errs.Locn, ErrParam): self.errs.RecordErr()
                return False
        return True

    def LstColsAllNonBlank(self, df=None, lst_cols=None):
        """
        .tbl.df list of tbl.nonblank_cols all contain at least one non-blank value (True if so)
        JDL 2/16/24; Modified 8/27/24 to add df arg
        """
        #list to check is tbl attribute unless overridden by args
        if lst_cols is None: lst_cols = self.tbl.nonblank_cols

        for col in lst_cols: 
            if not self.ColNonBlank(col, df=df,): return False
        return True

    def ColNonBlank(self, col_name, df=None):
        """
        Specified column contains no non-blank values (True if so)
        JDL 2/16/24; Modified 8/27/24 to add df arg
        """
        #Enable custom error codes and set df with precedence to arg df if supplied
        if not self.IsCustomCodes: self.errs.Locn = util.current_fn()
        if df is None: df = self.tbl.df

        #Check column contains at least one non-blank value
        is_col_nonblank = df[col_name].notnull().any()
        if self.errs.is_fail((not is_col_nonblank), 7, self.errs.Locn, col_name):
            self.errs.RecordErr()
            return False
        return True

    def LstColsAllNumeric(self, df=None, lst_cols=None):
        """
        .tbl list of .numeric_cols all numeric values (True if so)
        JDL 2/16/24; Modified 8/27/24 to add df arg
        """
        #list to check is tbl attribute unless overridden by args
        if lst_cols is None: lst_cols = self.tbl.numeric_cols

        for col in lst_cols:
            if not self.ColNumeric(col, df=df): return False
        return True

    def ColNumeric(self, col_name, df=None):
        """
        Values in a specified column are non-blank and numeric (True if so)
        JDL 2/19/24; Modified 8/27/24 to add df arg
        """
        #Enable custom error codes and set df with precedence to arg df if supplied
        if not self.IsCustomCodes: self.errs.Locn = util.current_fn()
        if df is None: df = self.tbl.df

        # Convert the column to numeric, coercing non-numeric values to NaN
        col_numeric = pd.to_numeric(df[col_name], errors='coerce')
        is_col_all_numeric = not col_numeric.isna().any()
        if self.errs.is_fail((not is_col_all_numeric), 8, self.errs.Locn, str(col_name)):
            self.errs.RecordErr()
            return False
        return True

    def LstColsAllInNumericRange(self, lst_cols, llim=None, ulim=None, df=None):
        """
        tbls.tbl1.df list of columns' values are within a specified numeric range
        JDL 2/19/24; Modified 8/27/24 to add df arg
        """
        for col in lst_cols:
            if not self.ColValsInNumericRange(col, llim=llim, ulim=ulim, df=df): return False
        return True

    def ColValsInNumericRange(self, col, llim=None, ulim=None, df=None):
        """
        Column values (must be numeric) are within specified range
        JDL 2/19/24; Modified 8/27/24 to add df arg
        """
        #Enable custom error codes and set df with precedence to arg df if supplied
        if not self.IsCustomCodes: self.errs.Locn = util.current_fn()
        if df is None: df = self.tbl.df

        is_in_range = True

        # If llim specified check column values greater than or equal to llim
        if llim is not None:
            if not df[col].ge(llim).all(): is_in_range = False

        # If ulim specified check column values less than or equal to ulim
        if ulim is not None:
            if not df[col].le(ulim).all(): is_in_range = False

        if self.errs.is_fail((not is_in_range), 9, self.errs.Locn, str(col)):
            self.errs.RecordErr()
            return False
        return True

    def ColValsMatchRegex(self, col_name, str_regex, IgnoreCase=False, df=None):
        """
        Column values match specified regex pattern
        JDL 3/1/24; Modified 8/27/24 to add df arg
        """
        #Enable custom error codes and set df with precedence to arg df if supplied
        if not self.IsCustomCodes: self.errs.Locn = util.current_fn()
        if df is None: df = self.tbl.df

        # Check if all column values match the regex
        fn_lambda = lambda x: bool(re.match(str_regex, str(x)))
        if IgnoreCase: fn_lambda = lambda x: bool(re.match(str_regex, str(x), re.IGNORECASE))
        is_match =  df[col_name].apply(fn_lambda).all()

        if self.errs.is_fail((not is_match), 10, self.errs.Locn, str(col_name)):
            self.errs.RecordErr()
            return False
        return True

    def ColContainsListVals(self, col_name, list_vals, df=None):
        """
        Individual column contains a specified list of values
        JDL 6/24/24; Modified 8/27/24 to add df arg
        """
        #Enable custom error codes and set df with precedence to arg df if supplied
        if not self.IsCustomCodes: self.errs.Locn = util.current_fn()
        if df is None: df = self.tbl.df

        # Loop over list_vals and check if in DataFrame index
        for val in list_vals:
            if val not in self.tbl.df[col_name].values:
                ErrParam = '\nMissing: ' + str(val)
                if self.errs.is_fail(True, 11, self.errs.Locn, ErrParam): self.errs.RecordErr()
                return False
        return True

    def ColContainsNodupsListVals(self, col_name, list_vals, df=None):
        """
        Column does not have duplicates of a list of values
        JDL 6/24/24; Modified 8/27/24 to add df arg
        """
        #Enable custom error codes and set df with precedence to arg df if supplied
        if not self.IsCustomCodes: self.errs.Locn = util.current_fn()
        if df is None: df = self.tbl.df

        # Loop over list_vals and check if in DataFrame index
        for val in list_vals:
            fil = df[col_name] == val
            if df.loc[fil, col_name].index.size > 1:
                ErrParam = '\nDuplicate: ' + str(val)
                if self.errs.is_fail(True, 12, self.errs.Locn, ErrParam): self.errs.RecordErr()
                return False
        return True

    def TableLocMatchesRegex(self, col_name1, val, col_name2, str_regex, \
                             IgnoreCase=False, df=None):
        """
        Specific table value matches regex pattern
        JDL 6/24/24; Modified 8/27/24 to add df arg
        """
        #Enable custom error codes and set df with precedence to arg df if supplied
        if not self.IsCustomCodes: self.errs.Locn = util.current_fn()
        if df is None: df = self.tbl.df

        # Filter DataFrame based on col_name1 and val to lookup col_name2 value
        fil = df[col_name1] == val
        val_loc = df.loc[fil, col_name2].values[0]

        # Compile regex with IGNORECASE flag if IgnoreCase is True
        if IgnoreCase:
            pattern = re.compile(str_regex, re.IGNORECASE)
        else:
            pattern = re.compile(str_regex)

        # Check if val_loc matches the regex pattern
        is_match = bool(pattern.match(val_loc))

        if self.errs.is_fail((not is_match), 13, self.errs.Locn):
            self.errs.ErrParam = '\nNon-match: ' + str(val_loc)
            self.errs.RecordErr()
            return False
        return True

    def NoDuplicateColVals(self, col, df=None):
        """
        Specified column does not have duplicate values (True if so)
        JDL 1/26/24
        """
        #Enable custom error codes and set df with precedence to arg df if supplied
        if not self.IsCustomCodes: self.errs.Locn = util.current_fn()
        if df is None: df = self.tbl.df

        if df[col].is_unique: return True
        if self.errs.is_fail(True, 14, self.errs.Locn, col): self.errs.RecordErr()
        return False
"""
=========================================================================
This class checks a specified list of Excel files (path+filename) for
existence and validity. For each file, it can also check that a specified
list of sheets exist within the Excel workbook.
=========================================================================
"""
class CheckExcelFiles:
    def __init__(self, lst_files, lst_shts, errs, IsErrsAsWarnings=True, IsPrint=False):
        """
        Initialize CheckExcelFiles with a list of file paths and a list of sheets.
        JDL 1/2/24
        """
        self.lst_files = lst_files # list of file paths to check
        self.lst_shts = lst_shts # list of lists of sheets

        self.wb = None # Current Excel workbook object
        self.IsWbErr = False # Flag if any errors during procedure

        # ErrorHandle instance and atts - Errors are treated as non-fatal warnings
        self.errs = errs
        self.errs.IsWarning = IsErrsAsWarnings
        self.errs.IsPrint = IsPrint

    def CheckFilesProcedure(self):
        """
        Procedure to check specified Excel files and sheets
        JDL 1/4/24
        """
        # Set location for looking up error messages
        self.errs.Locn = util.current_fn()

        for idx in range(len(self.lst_files)):

            # Check that file exists and can be opened; set self.wb object
            if not self.ExcelFileExists(idx): continue
            if not self.ExcelFileOpens(idx): continue

            # If valid Excel file, check if specified sheets are present
            if self.AllSheetsExist(idx): self.wb.Close(False)

    def ExcelFileExists(self, idx):
        """
        Check if an Excel file exists based on specified list index for list 
        of files to check (iteration over list in calling CheckFilesProcedure)
        JDL 1/4/24; Modified 2/7/24 to add shortening path for reporting
        """
        #If error, use .is_fail to set params including Locn of calling function
        fpath = self.lst_files[idx]

        if not os.path.exists(fpath):

            #Shorten the directory path for printing
            fpath = util.ck_for_shorten_path(self.lst_files[idx], 3)

            #Set errs params and report the error (add to self.errs.Msgs_Accum)
            self.errs.is_fail(True, 1, self.errs.Locn, '\n ' + fpath + '\n')
            self.IsWbErr = True
            self.errs.RecordErr()
            return False
        return True

    def ExcelFileOpens(self, idx):
        """
        Check if file is a valid Excel file based on ability to open
        JDL 1/4/24
        """
        fpath = self.lst_files[idx]
        try:
            self.wb = load_workbook(filename=fpath)
            return True
        
        #If error, use .is_fail to set params including Locn of calling function
        except Exception as e:
            self.IsWbErr = True

            #Shorten the directory path for printing
            fpath = util.ck_for_shorten_path(self.lst_files[idx], 3)

            #Record the error (add to self.errs.Msgs_Accum)
            self.errs.is_fail(True, 2, self.errs.Locn, '\n ' + fpath + '\n')
            self.errs.RecordErr()
            return False

    def AllSheetsExist(self, idx):
        """
        Iteratively check if specified sheets exist in self.wb
        JDL 1/4/24
        """
        for sheet_name in self.lst_shts[idx]:
            if not self.SheetExists(idx, sheet_name): break

    def SheetExists(self, idx, sheet_name):
        """
        Check if specified sheet exists in self.wb
        JDL 1/4/24
        """
        #Set the path and errParam string for current file and sheet
        fpath = self.lst_files[idx]

        #Set errParam string for current file and sheet (Shorten path for printing)
        fpath = util.ck_for_shorten_path(self.lst_files[idx], 3)
        errParam = '\n Missing: ' + sheet_name + ' in ' + fpath + '\n'

        #Check if sheet exists in self.wb
        is_sht_exists = sheet_name in self.wb.sheetnames
        if self.errs.is_fail(not is_sht_exists, 3, self.errs.Locn, errParam):
            self.IsWbErr = True
            self.errs.RecordErr()
            return False
        return True